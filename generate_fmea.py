import pandas as pd
import openpyxl
from copy import copy
import os
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime
from thefuzz import process


def find_official_phase(user_input, official_phases):
    choice = user_input.strip().lower()
    for phase in official_phases:
        if choice == str(phase).strip().lower():
            return phase
    if official_phases:
        best_match, score = process.extractOne(user_input, official_phases)
        if score >= 85:
            return best_match
    return user_input.strip()


def generate_fmea(template_path, masters_path, output_path, phases, title=None, rev="00", customer_only=False):
    if not os.path.exists(masters_path):
        raise FileNotFoundError(f"MASTERS Database not found: {masters_path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"FMEA Template not found: {template_path}")

    # FILTER PHASES IF IT IS CUSTOMER VERSION
    if customer_only:
        phases = [p for p in phases if p.get('is_customer', False)]
        if not phases:
            return  # If there are no customer phases, exit without creating the file

    df_master = pd.read_excel(masters_path, sheet_name='functional_database')
    df_master = df_master.dropna(how='all')
    phase_col = df_master.columns[0]
    master_cols_count = len(df_master.columns)

    official_db_phases = df_master[phase_col].dropna().astype(str).str.strip().unique().tolist()

    fmea_rows = []
    merge_blocks = []
    notes_to_append = []
    current_idx = 0

    for phase_info in phases:
        display_name_raw = phase_info['phase']
        input_phase_name = phase_info.get('base_phase', display_name_raw)
        db_phase_name = find_official_phase(input_phase_name, official_db_phases)
        op_num = phase_info['number']
        note = phase_info.get('note', '').strip()

        display_phase_name = display_name_raw.upper()
        if note:
            display_phase_name += " *"
            notes_to_append.append(f"* Op. {op_num} ({display_name_raw.upper()}): {note}")

        mask = df_master[phase_col].astype(str).str.strip().str.lower() == db_phase_name.strip().lower()
        matched = df_master[mask]

        if matched.empty:
            empty_row = [op_num, display_phase_name] + [""] * (master_cols_count - 1)
            fmea_rows.append(empty_row)
            block_size = 1
        else:
            block_size = len(matched)
            for _, row in matched.iterrows():
                row_list = ["" if pd.isna(x) else x for x in row.tolist()]
                row_list[0] = display_phase_name
                fmea_rows.append([op_num] + row_list)

        if block_size > 1:
            merge_blocks.append((current_idx, current_idx + block_size - 1))
        else:
            merge_blocks.append((current_idx, current_idx))
        current_idx += block_size

    normal_fmea_length = len(fmea_rows)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # RENAME SHEET BASED ON VERSION
    ws.title = "FMEA CUST" if customer_only else "FMEA"

    start_row = 5
    template_cols = ws.max_column

    if notes_to_append:
        fmea_rows.append([""] * template_cols)
        fmea_rows.append(["", "Additional Remarks:"] + [""] * (template_cols - 2))
        for n in notes_to_append:
            fmea_rows.append(["", n] + [""] * (template_cols - 2))

    today = datetime.now().strftime("%d/%m/%Y")
    try:
        ws['N3'].value = f"Date: {today}"
        ws['N3'].font = Font(name='Calibri Light', size=14, bold=True, scheme=None)
    except AttributeError:
        pass

    if title and title != "PRODUCT NAME":
        try:
            ws['C2'].value = title
            ws['C2'].font = Font(name='Calibri Light', size=14, bold=True, scheme=None)
            ws['C2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except AttributeError:
            pass

    try:
        ws['O2'].value = f"Rev: {rev}"
        ws['O2'].font = Font(name='Calibri Light', size=14, bold=True, scheme=None)
        ws['O2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except AttributeError:
        pass

    if len(fmea_rows) > 1:
        ws.insert_rows(6, amount=len(fmea_rows) - 1)

    for r_idx, row_data in enumerate(fmea_rows):
        excel_row = start_row + r_idx
        is_note = r_idx >= normal_fmea_length

        for c_idx in range(template_cols):
            col_num = c_idx + 1
            cell = ws.cell(row=excel_row, column=col_num)

            if type(cell).__name__ == 'MergedCell':
                continue

            if is_note:
                cell.value = row_data[c_idx] if c_idx < len(row_data) else ""
                is_bold_title = (c_idx == 1 and row_data[c_idx] == "Additional Remarks:")
                note_font_size = 14 if is_bold_title else 13
                cell.font = Font(name='Calibri Light', size=note_font_size, bold=is_bold_title, scheme=None)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, textRotation=0)
                continue

            if col_num == 11:
                cell.value = f'=E{excel_row}*H{excel_row}*J{excel_row}'
            elif col_num == 18:
                cell.value = f'=O{excel_row}*P{excel_row}*Q{excel_row}'
            elif col_num == 6:
                cell.value = f'=IF(E{excel_row}>6,"++",IF(E{excel_row}>3,"+",IF(E{excel_row}<4,"-","")))'
            elif c_idx < len(row_data):
                cell.value = row_data[c_idx]
            else:
                cell.value = ""

            if r_idx > 0:
                source_cell = ws.cell(row=5, column=col_num)
                if type(source_cell).__name__ != 'MergedCell':
                    cell.font = copy(source_cell.font)
                    cell.border = copy(source_cell.border)
                    cell.fill = copy(source_cell.fill)
                    cell.number_format = copy(source_cell.number_format)

            current_font = cell.font or Font()
            new_font = copy(current_font)
            new_font.bold = False
            new_font.name = 'Calibri Light'
            new_font.scheme = None  # Added theme decoupling!
            cell.font = new_font

            source_alignment = ws.cell(row=5, column=col_num).alignment
            new_align = copy(source_alignment) if source_alignment else Alignment()
            new_align.textRotation = 0
            new_align.wrap_text = True
            if col_num in [5, 6, 8, 10, 11, 14, 15, 16, 17, 18]:
                new_align.horizontal = 'center'
                new_align.vertical = 'center'
            cell.alignment = new_align

    columns_to_merge = [1, 2, 3, 4, 7, 9]
    default_side = Side(style='thin', color='000000')

    for start_idx, end_idx in merge_blocks:
        start_excel_row = start_row + start_idx
        end_excel_row = start_row + end_idx

        for r in range(start_excel_row, end_excel_row + 1):
            for c in range(1, template_cols + 1):
                cell = ws.cell(row=r, column=c)
                if type(cell).__name__ == 'MergedCell':
                    continue

                cb = cell.border or Border()
                new_left = default_side
                new_right = default_side
                new_top = default_side if r == start_excel_row else cb.top
                new_bottom = default_side if r == end_excel_row else cb.bottom

                cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom,
                                     diagonal=cb.diagonal, diagonal_direction=cb.diagonal_direction)

        for col in columns_to_merge:
            sub_start = start_excel_row
            while sub_start <= end_excel_row:
                val = ws.cell(row=sub_start, column=col).value
                sub_end = sub_start
                while sub_end < end_excel_row and ws.cell(row=sub_end + 1, column=col).value == val:
                    sub_end += 1

                if sub_end > sub_start:
                    ws.merge_cells(start_row=sub_start, start_column=col, end_row=sub_end, end_column=col)
                    top_cell = ws.cell(row=sub_start, column=col)
                    if type(top_cell).__name__ != 'MergedCell':
                        orig_align = top_cell.alignment
                        top_cell.alignment = Alignment(horizontal=orig_align.horizontal if orig_align else 'left',
                                                       vertical='center', wrap_text=True, textRotation=0)
                sub_start = sub_end + 1

    if notes_to_append:
        note_start_row = start_row + normal_fmea_length + 1
        for r in range(note_start_row, note_start_row + len(notes_to_append) + 1):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=template_cols)

    # --- GLOBAL FONT FORCING: FINAL ANTI-APTOS BLOCK ---
    for row in ws.iter_rows():
        for cell in row:
            if type(cell).__name__ == 'MergedCell':
                continue
            if cell.font:
                new_font = copy(cell.font)
                new_font.name = 'Calibri Light'
                new_font.scheme = None  # Total removal of theme inheritance
                cell.font = new_font
            else:
                cell.font = Font(name='Calibri Light', scheme=None)
    # ---------------------------------------------------------

    ws.freeze_panes = None
    ws.sheet_view.topLeftCell = 'A1'
    ws.freeze_panes = 'A5'
    ws.print_title_rows = '1:4'

    wb.save(output_path)
    print(f"✅ FMEA generated in: {output_path}")
