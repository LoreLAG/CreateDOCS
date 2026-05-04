import pandas as pd
import openpyxl
from copy import deepcopy, copy
import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime
from thefuzz import process, fuzz  # Importato fuzz per il punteggio incrociato


def find_best_matches(phase_info, fasi_pure_db, blocchi_db):
    """
    Trova i suggerimenti e calcola il miglior abbinamento assoluto
    basandosi su Fase, Reparto e penalizzando incongruenze Int/Ext.
    """
    fase_gui = str(phase_info['base_phase']).strip().upper()
    dept_gui = str(phase_info.get('department', '')).strip().upper()
    ext_int_gui = str(phase_info.get('ext_int', '')).strip().upper()

    # 1. Trova i candidati suggeriti in base al NOME DELLA FASE
    match_uppers = [f_db.upper() for f_db in fasi_pure_db if fase_gui in f_db.upper()]
    if not match_uppers and fasi_pure_db:
        matches = process.extract(fase_gui, fasi_pure_db, limit=15)
        match_uppers = [m[0].upper() for m in matches if m[1] >= 60]

    match_uppers = list(dict.fromkeys(match_uppers))
    suggested = [k for k in blocchi_db.keys() if k[0] in match_uppers]

    # 2. Calcola il miglior abbinamento tenendo conto di REPARTO e INT/EXT
    best_key = None
    best_score = -999

    is_internal_gui = ext_int_gui in ["INT", "INTERNAL", "INT.", "I"]
    is_external_gui = ext_int_gui in ["EXT", "EXTERNAL", "EXT.", "E"]

    for k in suggested:
        f_db, t_db = k[0], k[1]

        # Punteggi base
        score_fase = fuzz.WRatio(fase_gui, f_db)
        score_dept = fuzz.WRatio(dept_gui, t_db) if dept_gui else 0

        # Logica di Penalità Int/Ext
        penalty = 0
        if is_internal_gui and ("EXT" in t_db or "EXTERNAL" in t_db):
            penalty += 50
        if is_external_gui and ("INT" in t_db or "INTERNAL" in t_db):
            penalty += 50

        # Calcolo Totale (Fase conta al 60%, Reparto al 40%, meno la penalità)
        total_score = (score_fase * 0.6) + (score_dept * 0.4) - penalty

        if total_score > best_score:
            best_score = total_score
            best_key = k

    return suggested, best_key


def chiedi_scelte_globali(phases, fasi_pure_db, blocchi_db, scelte_cache):
    """
    Apre un'interfaccia a 3 colonne:
    SX: Database | CENTRO: Composizione Fase Corrente | DX: Segnalibri Fasi
    """
    scelta_finale = {"abort": True, "scelte": {}}

    parent = tk._default_root
    if parent:
        fw = parent.focus_get()
        if fw: parent = fw.winfo_toplevel()

    root = tk.Toplevel(parent)
    if parent: root.transient(parent)
    root.withdraw()
    root.title("Compositore Globale Control Plan")

    # ─── STATO GLOBALE DELL'INTERFACCIA ───
    state = {
        "active_idx": 0,
        "scelte": {},
        "current_chiave": None,
        "current_suggested_keys": [],
        "visited": set()
    }

    # Pre-calcola gli abbinamenti per tutte le fasi con la nuova logica intelligente
    for i, p in enumerate(phases):
        fase_gui = str(p['base_phase']).strip()
        chiave = f"{p['number']}_{fase_gui}"

        if chiave in scelte_cache:
            state["scelte"][chiave] = deepcopy(scelte_cache[chiave])
            state["visited"].add(i)
        else:
            suggested, best_key = find_best_matches(p, fasi_pure_db, blocchi_db)

            if best_key:
                state["scelte"][chiave] = [{"key": best_key, "alias": None}]
            else:
                state["scelte"][chiave] = []

    # ─── LAYOUT PRINCIPALE ───
    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # FRAME DESTRA: Segnalibri Fasi
    nav_frame = ttk.LabelFrame(main_frame, text="Fasi (Usa frecce Su/Giù)")
    nav_frame.pack(side="right", fill="y", padx=(10, 0))

    scroll_nav = ttk.Scrollbar(nav_frame)
    scroll_nav.pack(side="right", fill="y")
    listbox_nav = tk.Listbox(nav_frame, selectmode=tk.SINGLE, yscrollcommand=scroll_nav.set, width=45,
                             font=("Segoe UI", 10))
    listbox_nav.pack(side="left", fill="both", expand=True, padx=5, pady=5)
    scroll_nav.config(command=listbox_nav.yview)

    # FRAME SINISTRA E CENTRO
    editor_frame = tk.Frame(main_frame)
    editor_frame.pack(side="left", fill="both", expand=True)

    lbl_fase_corrente = ttk.Label(editor_frame, text="", font=("Segoe UI", 12, "bold"), foreground="#005599")
    lbl_fase_corrente.pack(anchor="w", pady=(0, 10))

    columns_frame = tk.Frame(editor_frame)
    columns_frame.pack(fill="both", expand=True)

    columns_frame.columnconfigure(0, weight=1)
    columns_frame.columnconfigure(1, weight=0)
    columns_frame.columnconfigure(2, weight=1)
    columns_frame.columnconfigure(3, weight=0)
    columns_frame.rowconfigure(2, weight=1)

    # --- RIGA 0: Intestazioni ---
    top_db_frame = tk.Frame(columns_frame)
    top_db_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
    lbl_titolo_sx = ttk.Label(top_db_frame, text="Varianti suggerite:", font=("Segoe UI", 9, "bold"))
    lbl_titolo_sx.pack(side="left")

    mostra_tutto_var = tk.BooleanVar(value=False)
    cb_mostra_tutto = tk.Checkbutton(top_db_frame, text="Mostra tutto", variable=mostra_tutto_var)
    cb_mostra_tutto.pack(side="right")

    top_cp_frame = tk.Frame(columns_frame)
    top_cp_frame.grid(row=0, column=2, sticky="ew", pady=(0, 5))
    ttk.Label(top_cp_frame, text="Blocchi per questa fase (Tasto DX per Alias):", font=("Segoe UI", 9, "bold")).pack(
        side="left")

    # --- RIGA 1: Filtro ---
    filter_var = tk.StringVar()
    filter_entry = ttk.Entry(columns_frame, textvariable=filter_var)
    filter_entry.grid(row=1, column=0, sticky="ew", pady=(0, 5))
    filter_entry.insert(0, "Cerca fase...")

    def on_filter_in(e):
        if filter_var.get() == "Cerca fase...": filter_entry.delete(0, tk.END)

    def on_filter_out(e):
        if not filter_var.get(): filter_entry.insert(0, "Cerca fase...")

    filter_entry.bind("<FocusIn>", on_filter_in)
    filter_entry.bind("<FocusOut>", on_filter_out)

    # --- RIGA 2: Listboxes ---
    list_frame_db = tk.Frame(columns_frame)
    list_frame_db.grid(row=2, column=0, sticky="nsew")
    scroll_db = ttk.Scrollbar(list_frame_db)
    scroll_db.pack(side="right", fill="y")
    listbox_db = tk.Listbox(list_frame_db, selectmode=tk.EXTENDED, yscrollcommand=scroll_db.set, width=55)
    listbox_db.pack(side="left", fill="both", expand=True)
    scroll_db.config(command=listbox_db.yview)

    list_frame_cp = tk.Frame(columns_frame)
    list_frame_cp.grid(row=2, column=2, sticky="nsew")
    scroll_cp = ttk.Scrollbar(list_frame_cp)
    scroll_cp.pack(side="right", fill="y")
    listbox_cp = tk.Listbox(list_frame_cp, selectmode=tk.SINGLE, yscrollcommand=scroll_cp.set, width=65)
    listbox_cp.pack(side="left", fill="both", expand=True)
    scroll_cp.config(command=listbox_cp.yview)

    # --- PULSANTI CENTRALI E LATERALI ---
    btn_frame = tk.Frame(columns_frame)
    btn_frame.grid(row=2, column=1, sticky="n", padx=10, pady=40)
    ttk.Button(btn_frame, text="Aggiungi DB >>", command=lambda: add_selected()).pack(pady=5)
    ttk.Button(btn_frame, text="Aggiungi Vuota >>", command=lambda: add_empty()).pack(pady=5)

    ord_frame = tk.Frame(columns_frame)
    ord_frame.grid(row=2, column=3, sticky="n", padx=5, pady=40)
    ttk.Button(ord_frame, text="⬆", width=3, command=lambda: move_up()).pack(pady=2)
    ttk.Button(ord_frame, text="⬇", width=3, command=lambda: move_down()).pack(pady=2)
    ttk.Button(ord_frame, text="❌", width=3, command=lambda: remove_sel()).pack(pady=(15, 2))

    # ─── FUNZIONI DI SUPPORTO ───
    def get_display_text(item_dict):
        key = item_dict["key"]
        alias = item_dict.get("alias")
        base_text = "[VUOTA] Riga vuota" if key == "VUOTO" else f"[DB] {key[0]} - {key[1]}"
        if alias:
            return f"[ALIAS: {alias}] (Orig: {key[0] if key != 'VUOTO' else 'Vuota'})"
        return base_text

    def render_nav_list():
        listbox_nav.delete(0, tk.END)
        for i, p in enumerate(phases):
            chiave = f"{p['number']}_{str(p['base_phase']).strip()}"
            has_items = len(state["scelte"].get(chiave, [])) > 0

            if not has_items:
                status = "❌"
                color = "red"
            elif i not in state["visited"]:
                status = "❗"
                color = "#b85c00"
            else:
                status = "✅"
                color = "#008000"

            fase_disp = p.get('phase', p['base_phase'])
            testo = f"{status} Op.{p['number']} - {fase_disp}"
            listbox_nav.insert(tk.END, testo)

            if color != "black":
                listbox_nav.itemconfig(i, {'fg': color})

        listbox_nav.selection_clear(0, tk.END)
        listbox_nav.selection_set(state["active_idx"])
        listbox_nav.see(state["active_idx"])

    def apply_db_filter(*args):
        q = filter_var.get().lower()
        if q == "cerca fase...": q = ""
        listbox_db.delete(0, tk.END)

        base_dict = blocchi_db if mostra_tutto_var.get() else state["current_suggested_keys"]
        base_keys = list(base_dict.keys()) if isinstance(base_dict, dict) else base_dict

        current_view_keys = [k for k in base_keys if not q or q in k[0].lower() or q in k[1].lower()]
        listbox_db.current_keys = current_view_keys

        for k in current_view_keys:
            listbox_db.insert(tk.END, f"{k[0]} - {k[1]}")

        if not current_view_keys:
            listbox_db.insert(tk.END, "Nessun risultato.")
            listbox_db.config(state=tk.DISABLED)
        else:
            listbox_db.config(state=tk.NORMAL)

    filter_var.trace_add("write", apply_db_filter)

    def load_phase(idx):
        if idx < 0 or idx >= len(phases): return

        state["active_idx"] = idx
        state["visited"].add(idx)

        p = phases[idx]
        fase_gui = str(p['base_phase']).strip()
        state["current_chiave"] = f"{p['number']}_{fase_gui}"

        fase_disp = p.get('phase', fase_gui)
        dept_disp = p.get('department', 'N/D')
        lbl_fase_corrente.config(text=f"⚙️ In modifica: Op. {p['number']} - {fase_disp} [{dept_disp}]")

        suggested, _ = find_best_matches(p, fasi_pure_db, blocchi_db)
        state["current_suggested_keys"] = suggested

        if not suggested:
            mostra_tutto_var.set(True)

        lbl_titolo_sx.config(text="Tutte le varianti (DB):" if mostra_tutto_var.get() else "Varianti suggerite:")

        if filter_var.get() != "Cerca fase...":
            filter_var.set("")
        apply_db_filter()

        listbox_cp.delete(0, tk.END)
        for item in state["scelte"][state["current_chiave"]]:
            listbox_cp.insert(tk.END, get_display_text(item))

        render_nav_list()

    def on_nav_click(event):
        sel = listbox_nav.curselection()
        if sel:
            load_phase(sel[0])

    listbox_nav.bind("<<ListboxSelect>>", on_nav_click)
    cb_mostra_tutto.config(command=lambda: load_phase(state["active_idx"]))

    # ─── AZIONI EDITING E TASTIERA ───
    def nav_up(e):
        load_phase(state["active_idx"] - 1)
        return "break"

    def nav_down(e):
        load_phase(state["active_idx"] + 1)
        return "break"

    def global_nav_up(e):
        if e.widget in (listbox_db, listbox_cp, filter_entry): return
        nav_up(e)

    def global_nav_down(e):
        if e.widget in (listbox_db, listbox_cp, filter_entry): return
        nav_down(e)

    listbox_nav.bind("<Up>", nav_up)
    listbox_nav.bind("<Down>", nav_down)
    root.bind("<Up>", global_nav_up)
    root.bind("<Down>", global_nav_down)

    def add_selected(event=None):
        if not getattr(listbox_db, 'current_keys', None): return
        sel = listbox_db.curselection()
        for i in sel:
            item = {"key": listbox_db.current_keys[i], "alias": None}
            state["scelte"][state["current_chiave"]].append(item)
            listbox_cp.insert(tk.END, get_display_text(item))
        listbox_db.selection_clear(0, tk.END)
        render_nav_list()

    def add_empty():
        item = {"key": "VUOTO", "alias": None}
        state["scelte"][state["current_chiave"]].append(item)
        listbox_cp.insert(tk.END, get_display_text(item))
        render_nav_list()

    def set_alias(event):
        idx = listbox_cp.nearest(event.y)
        if idx < 0 or idx >= listbox_cp.size(): return
        bbox = listbox_cp.bbox(idx)
        if not bbox or not (bbox[1] <= event.y <= bbox[1] + bbox[3]): return

        listbox_cp.selection_clear(0, tk.END)
        listbox_cp.selection_set(idx)

        item = state["scelte"][state["current_chiave"]][idx]
        current_alias = item.get("alias", "") or ""

        new_alias = simpledialog.askstring(
            "Imposta Alias",
            "Inserisci il nome che vuoi far comparire sul Control Plan:",
            initialvalue=current_alias, parent=root
        )

        if new_alias is not None:
            item["alias"] = new_alias.strip() if new_alias.strip() else None
            listbox_cp.delete(idx)
            listbox_cp.insert(idx, get_display_text(item))
            listbox_cp.selection_set(idx)

    def remove_sel(event=None):
        sel = listbox_cp.curselection()
        if not sel: return
        idx = sel[0]
        listbox_cp.delete(idx)
        state["scelte"][state["current_chiave"]].pop(idx)
        if listbox_cp.size() > 0:
            listbox_cp.selection_set(max(0, idx - 1))
        render_nav_list()

    def move_up():
        sel = listbox_cp.curselection()
        if not sel: return
        idx = sel[0]
        if idx > 0:
            lst = state["scelte"][state["current_chiave"]]
            lst[idx - 1], lst[idx] = lst[idx], lst[idx - 1]
            val = listbox_cp.get(idx)
            listbox_cp.delete(idx)
            listbox_cp.insert(idx - 1, val)
            listbox_cp.selection_set(idx - 1)

    def move_down():
        sel = listbox_cp.curselection()
        if not sel: return
        idx = sel[0]
        lst = state["scelte"][state["current_chiave"]]
        if idx < len(lst) - 1:
            lst[idx + 1], lst[idx] = lst[idx], lst[idx + 1]
            val = listbox_cp.get(idx)
            listbox_cp.delete(idx)
            listbox_cp.insert(idx + 1, val)
            listbox_cp.selection_set(idx + 1)

    listbox_db.bind("<Double-1>", add_selected)
    listbox_cp.bind("<Double-1>", remove_sel)
    listbox_cp.bind("<Button-3>", set_alias)

    # ─── BOTTOM: CONFERMA GLOBALE ───
    bot_frame = tk.Frame(root)
    bot_frame.pack(fill="x", pady=10, padx=10)

    def conferma_tutto(event=None):
        vuote = [p['number'] for p in phases if not state["scelte"][f"{p['number']}_{str(p['base_phase']).strip()}"]]
        if vuote:
            msg = f"Attenzione: le seguenti operazioni non hanno blocchi assegnati:\nOp. {', '.join(map(str, vuote))}\n\nVuoi riempire queste fasi con righe vuote automaticamente?"
            ans = messagebox.askyesnocancel("Fasi incomplete", msg, parent=root)
            if ans is None: return
            if ans is True:
                for v in vuote:
                    for p in phases:
                        if p['number'] == v:
                            ch = f"{p['number']}_{str(p['base_phase']).strip()}"
                            state["scelte"][ch].append({"key": "VUOTO", "alias": None})
                            break

        scelta_finale["abort"] = False
        scelta_finale["scelte"] = state["scelte"]
        root.destroy()

    ttk.Button(bot_frame, text="Salva e Genera CP", command=conferma_tutto, style="Accent.TButton").pack()
    root.bind("<Return>", conferma_tutto)

    if phases:
        load_phase(0)

    root.update_idletasks()
    window_width = 1250
    window_height = 650
    try:
        if parent and parent.winfo_viewable():
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            pos_x = px + (pw // 2) - (window_width // 2)
            pos_y = py + (ph // 2) - (window_height // 2)
        else:
            sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
            pos_x = (sw // 2) - (window_width // 2)
            pos_y = (sh // 2) - (window_height // 2)
    except:
        pos_x, pos_y = 100, 100

    root.geometry(f"{window_width}x{window_height}+{pos_x}+{pos_y}")
    root.deiconify()
    root.grab_set()
    root.focus_force()
    root.wait_window()

    return None if scelta_finale["abort"] else scelta_finale["scelte"]


def generate_cp(template_path, masters_path, output_path, phases, title=None, rev="00", customer_only=False,
                scelte_cache=None):
    if not os.path.exists(masters_path):
        raise FileNotFoundError(f"Database MASTERS non trovato: {masters_path}")

    if customer_only:
        phases = [p for p in phases if p.get('is_customer', False)]
        if not phases:
            return

    df = pd.read_excel(masters_path, sheet_name='database_funzionale')
    df = df.dropna(how='all')
    col_fase = df.columns[0]
    col_tipo = df.columns[1]

    fasi_pure_db = df[col_fase].dropna().astype(str).str.strip().unique().tolist()

    blocchi_db = {}
    for (f, t), group in df.groupby([col_fase, col_tipo], sort=False):
        blocchi_db[(str(f).strip().upper(), str(t).strip().upper())] = group

    if scelte_cache is None:
        scelte_cache = {}

    # APERTURA COMPOSITORE GLOBALE
    nuove_scelte = chiedi_scelte_globali(phases, fasi_pure_db, blocchi_db, scelte_cache)
    if nuove_scelte is None:
        raise InterruptedError("Generazione annullata dall'utente nel Control Plan.")

    # Aggiorna la cache con le modifiche globali
    scelte_cache.update(nuove_scelte)

    start_row = 11
    cp_rows_main = []
    merge_info_main = []
    current_row_main = start_row

    for phase_info in phases:
        fase_gui = str(phase_info['base_phase']).strip()
        op_num = phase_info['number']
        fase_display = phase_info.get('phase', fase_gui)

        chiave_fase = f"{op_num}_{fase_gui}"
        lista_scelte = scelte_cache.get(chiave_fase, [])

        for scelta_dict in lista_scelte:
            # Compatibilità per sicurezza
            if isinstance(scelta_dict, dict):
                scelta_item = scelta_dict["key"]
                alias = scelta_dict.get("alias")
            else:
                scelta_item = scelta_dict
                alias = None

            if scelta_item == "VUOTO":
                blocco_scelto_df = pd.DataFrame(
                    [["", phase_info.get('ext_int', '')] + [""] * (len(df.columns) - 2)], columns=df.columns)
            else:
                blocco_scelto_df = blocchi_db.get(scelta_item)
                if blocco_scelto_df is None:
                    blocco_scelto_df = pd.DataFrame(
                        [["", phase_info.get('ext_int', '')] + [""] * (len(df.columns) - 2)], columns=df.columns)

            num_righe_blocco = len(blocco_scelto_df)
            merge_info_main.append((current_row_main, current_row_main + num_righe_blocco - 1))
            current_row_main += num_righe_blocco

            for _, row in blocco_scelto_df.iterrows():
                row_list = ["" if pd.isna(x) else x for x in row.tolist()]

                # Applica l'alias o il nome della GUI
                nome_da_stampare = alias.upper() if alias else fase_display.upper()
                row_list[0] = nome_da_stampare

                final_row = [op_num] + row_list
                cp_rows_main.append(final_row)

    wb = openpyxl.load_workbook(template_path)
    ws_main = wb.active
    ws_main.title = "Control Plan CUST" if customer_only else "Control Plan"

    if len(cp_rows_main) > 1:
        ws_main.insert_rows(start_row + 1, amount=len(cp_rows_main) - 1)

    oggi = datetime.now().strftime("%d/%m/%Y")
    ws_main['J7'] = f"Date: {oggi}"
    ws_main['P7'] = f"Rev: {rev}"
    if title: ws_main['C5'] = title

    thin = Side(style='thin', color="000000")
    for r_idx, row_data in enumerate(cp_rows_main):
        r_excel = start_row + r_idx
        ws_main.row_dimensions[r_excel].height = 23
        for c_idx, val in enumerate(row_data):
            c_excel = c_idx + 1
            if c_excel > ws_main.max_column: break

            cell = ws_main.cell(row=r_excel, column=c_excel)
            cell.value = val

            if cell.font:
                new_font = copy(cell.font)
                new_font.name = 'Calibri Light'
                new_font.scheme = None
                cell.font = new_font
            else:
                cell.font = Font(name='Calibri Light', size=11, scheme=None)

            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # UNIONE CELLE PER BLOCCO
    for r_start, r_end in merge_info_main:
        if r_start != r_end:
            for col in [1, 2, 3]:
                ws_main.merge_cells(start_row=r_start, start_column=col, end_row=r_end, end_column=col)
                top_cell = ws_main.cell(row=r_start, column=col)
                top_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # RIPULITURA INTESTAZIONE E CORPO
    for row in ws_main.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if type(cell).__name__ == 'MergedCell':
                continue
            if cell.font:
                new_font = copy(cell.font)
                new_font.name = 'Calibri Light'
                new_font.scheme = None
                cell.font = new_font
            else:
                cell.font = Font(name='Calibri Light', scheme=None)

    ws_main.sheet_view.topLeftCell = 'A1'
    ws_main.freeze_panes = 'A11'

    wb.save(output_path)
    print(f"Control Plan generato in: {output_path}")