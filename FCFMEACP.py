import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import ctypes
import shutil
import time
import json
import zipfile
import re
import xml.etree.ElementTree as ET
import openpyxl

from generate_flowchart import generate as generate_fc
from generate_fmea import generate_fmea
from generate_cp import generate_cp
from LeggiMasterFMEA import format_masters_db
from LeggiMasterCP import format_master_cp_db

# Necessario per mostrare l'icona personalizzata nella taskbar di Windows
try:
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("BoringApp")
except:
    pass


def resource_path(relative_path):
    """Ottiene il percorso assoluto delle risorse incapsulate in PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR_NAME = "File utili per FC-FMEA-CP"
DATA_DIR = os.path.join(APP_DIR, DATA_DIR_NAME)

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)
    if sys.platform == 'win32':
        try:
            FILE_ATTRIBUTE_HIDDEN = 0x02
            ctypes.windll.kernel32.SetFileAttributesW(DATA_DIR, FILE_ATTRIBUTE_HIDDEN)
        except:
            pass

FILES_TO_EXTRACT = [
    "MASTERS-FMEA.xlsx", "MASTERS-CP.xlsx", "TEMPLATE_FLOW_CHART.xlsm",
    "TEMPLATE_FMEA.xlsx", "TEMPLATE_CP.xlsx", "MANUALE CREA DOCS.pdf"
]

for filename in FILES_TO_EXTRACT:
    local_path = os.path.join(DATA_DIR, filename)
    bundled_path = resource_path(filename)
    if not os.path.exists(local_path) and os.path.abspath(local_path) != os.path.abspath(bundled_path):
        try:
            shutil.copy2(bundled_path, local_path)
        except:
            pass

CONFIG_FILE = os.path.join(DATA_DIR, "config.json")


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_config(config_dict):
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config_dict, f)
    except:
        pass


def get_smart_path(filename, config_paths):
    key_map = {
        "TEMPLATE_FLOW_CHART.xlsm": "tpl_fc", "TEMPLATE_FMEA.xlsx": "tpl_fmea",
        "TEMPLATE_CP.xlsx": "tpl_cp", "MASTERS-FMEA.xlsx": "db_master",
        "MASTERS-CP.xlsx": "db_master_cp"
    }
    config_key = key_map.get(filename)
    if config_key and config_key in config_paths:
        custom_path = config_paths[config_key]
        if os.path.exists(custom_path): return custom_path
    default_path = os.path.join(DATA_DIR, filename)
    if os.path.exists(default_path): return default_path
    return resource_path(filename)


def format_short_path(full_path):
    if not full_path: return ""
    parts = full_path.replace('\\', '/').split('/')
    return ".../" + "/".join(parts[-3:]) if len(parts) > 3 else full_path


class FlowChartApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generatore Produzione (FC, FMEA & CP)")
        window_width, window_height = 900, 800
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        self.root.geometry(
            f"{window_width}x{window_height}+{int(sw / 2 - window_width / 2)}+{int(sh / 2 - window_height / 2)}")

        self.phases_data = []
        self.saved_config = load_config()
        self.real_paths = {
            "tpl_fc": get_smart_path("TEMPLATE_FLOW_CHART.xlsm", self.saved_config),
            "tpl_fmea": get_smart_path("TEMPLATE_FMEA.xlsx", self.saved_config),
            "tpl_cp": get_smart_path("TEMPLATE_CP.xlsx", self.saved_config),
            "db_master": get_smart_path("MASTERS-FMEA.xlsx", self.saved_config),
            "db_master_cp": get_smart_path("MASTERS-CP.xlsx", self.saved_config)
        }

        self.tpl_fc_var = tk.StringVar(value=format_short_path(self.real_paths["tpl_fc"]))
        self.tpl_fmea_var = tk.StringVar(value=format_short_path(self.real_paths["tpl_fmea"]))
        self.tpl_cp_var = tk.StringVar(value=format_short_path(self.real_paths["tpl_cp"]))
        self.db_master_var = tk.StringVar(value=format_short_path(self.real_paths["db_master"]))
        self.db_master_var_cp = tk.StringVar(value=format_short_path(self.real_paths["db_master_cp"]))

        self.archive = self._load_archive_from_excel()

        # UI
        file_frame = ttk.LabelFrame(root, text="File di Sistema e Template", padding=(10, 5))
        file_frame.pack(fill="x", padx=10, pady=(10, 0))
        self._add_file_row(file_frame, "Template Flow Chart:", self.tpl_fc_var, "tpl_fc", ".xlsm", 0)
        self._add_file_row(file_frame, "Template FMEA:", self.tpl_fmea_var, "tpl_fmea", ".xlsx", 1)
        self._add_file_row(file_frame, "Template Control Plan:", self.tpl_cp_var, "tpl_cp", ".xlsx", 2)

        # Riga DB FMEA
        ttk.Label(file_frame, text="Database MASTERS FMEA:").grid(row=3, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(file_frame, textvariable=self.db_master_var, state="readonly", width=80).grid(row=3, column=1,
                                                                                                sticky="we", padx=5,
                                                                                                pady=2)
        db_btn_frame = tk.Frame(file_frame)
        db_btn_frame.grid(row=3, column=2, sticky="w")
        ttk.Button(db_btn_frame, text="Sfoglia...", command=self._choose_db_file).pack(side="left", padx=(5, 2))
        ttk.Button(db_btn_frame, text="✨ Pulisci DB", command=self.format_master_db).pack(side="left", padx=(2, 5))

        # Riga DB CP
        ttk.Label(file_frame, text="Database MASTERS CP:").grid(row=4, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(file_frame, textvariable=self.db_master_var_cp, state="readonly", width=48).grid(row=4, column=1,
                                                                                                   sticky="we", padx=5,
                                                                                                   pady=2)
        db_cp_btn_frame = tk.Frame(file_frame)
        db_cp_btn_frame.grid(row=4, column=2, sticky="w")
        ttk.Button(db_cp_btn_frame, text="Sfoglia...", command=self._choose_db_cp_file).pack(side="left", padx=(5, 2))
        ttk.Button(db_cp_btn_frame, text="✨ Pulisci DB", command=self.format_master_cp_db_gui).pack(side="left",
                                                                                                    padx=(2, 5))

        # Input Fasi
        input_frame = ttk.LabelFrame(root, text="Aggiungi una nuova fase", padding=(10, 10))
        input_frame.pack(fill="x", padx=10, pady=(10, 4))
        self.num_var = tk.StringVar(value="10")
        ttk.Label(input_frame, text="N° Op.:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        ttk.Entry(input_frame, textvariable=self.num_var, width=8).grid(row=0, column=1, sticky="w", padx=5, pady=5)
        self.phase_var = tk.StringVar()
        self.phase_combo = ttk.Combobox(input_frame, textvariable=self.phase_var, width=48)
        self.phase_combo.grid(row=0, column=3, sticky="w", padx=5, pady=5)
        self.phase_combo.bind("<<ComboboxSelected>>", self._on_phase_select)
        self.phase_combo.bind("<FocusOut>", self._on_phase_select)
        ttk.Label(input_frame, text="Nome Fase:").grid(row=0, column=2, sticky="e", padx=5, pady=5)
        self.extint_var = tk.StringVar(value="Internal")
        ttk.Combobox(input_frame, textvariable=self.extint_var, values=["Internal", "External"], width=10,
                     state="readonly").grid(row=1, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(input_frame, text="Int / Ext:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.dept_var = tk.StringVar()
        self.dept_combo = ttk.Combobox(input_frame, textvariable=self.dept_var, width=48)
        self.dept_combo.grid(row=1, column=3, sticky="w", padx=5, pady=5)
        ttk.Label(input_frame, text="Reparto:").grid(row=1, column=2, sticky="e", padx=5, pady=5)
        self.comp_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.comp_var, width=13).grid(row=2, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(input_frame, text="Componente:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.note_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.note_var, width=50).grid(row=2, column=3, sticky="w", padx=5, pady=5)
        ttk.Label(input_frame, text="Additional remarks:").grid(row=2, column=2, sticky="e", padx=5, pady=5)

        btn_frame_top = tk.Frame(input_frame)
        btn_frame_top.grid(row=3, column=0, columnspan=4, pady=(10, 2))
        ttk.Button(btn_frame_top, text="💾 Salva nel Foglio 'Fasi'", command=self._save_to_archive).pack(side="left",
                                                                                                        padx=5)
        ttk.Button(btn_frame_top, text="🗑Elimina dal Foglio 'Fasi'", command=self._delete_from_archive).pack(
            side="left", padx=5)
        ttk.Button(btn_frame_top, text="➕ AGGIUNGI FASE AL PROGETTO", command=self.add_phase,
                   style="Accent.TButton").pack(side="right", padx=15)

        self._refresh_archive_combo()
        self.phase_combo.bind("<KeyRelease>",
                              lambda e: self._autocomplete(e, self.phase_combo, [p["phase"] for p in self.archive]))
        self.dept_combo.bind("<KeyRelease>", lambda e: self._autocomplete(e, self.dept_combo, self._get_unique_depts()))
        self.root.bind("<Return>", self.add_phase)

        # Tabella
        list_frame = ttk.LabelFrame(root, text="Fasi di Produzione in Coda", padding=(10, 10))
        list_frame.pack(fill="both", expand=True, padx=10, pady=4)
        columns = ("Num", "Phase", "ExtInt", "Dept", "Note", "Cust")

        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=8, selectmode="extended")
        for col, head in zip(columns,
                             ["N° Op.", "Fase (Componente)", "Int/Ext", "Reparto", "Additional remarks", "Cliente?"]):
            self.tree.heading(col, text=head)
        self.tree.column("Num", width=50, anchor="center")
        self.tree.column("Phase", width=220)
        self.tree.column("ExtInt", width=60, anchor="center")
        self.tree.column("Cust", width=70, anchor="center")
        self.tree.pack(fill="both", expand=True, side="left")
        self.tree.bind("<Double-1>", self._on_double_click)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        action_frame = tk.Frame(root)
        action_frame.pack(fill="x", padx=10, pady=4)
        ttk.Button(action_frame, text="📥 Importa da FC / MASTER", command=self.import_from_fc).pack(side="left", padx=5)
        ttk.Button(action_frame, text="⬆️ Sposta Su", command=self.move_up).pack(side="left", padx=25)
        ttk.Button(action_frame, text="⬇️ Sposta Giù", command=self.move_down).pack(side="left", padx=5)

        ttk.Button(action_frame, text="🧹 Pulisci Tabella", command=self.clear_table).pack(side="right", padx=5)
        ttk.Button(action_frame, text="❌ Rimuovi Selezionate", command=self.remove_phase).pack(side="right", padx=5)

        # Generazione
        gen_frame = tk.Frame(root)
        gen_frame.pack(fill="x", padx=10, pady=12)
        self.title_var = tk.StringVar(value="NOME PRODOTTO")
        ttk.Label(gen_frame, text="Nome Prodotto:").pack(side="left")
        ttk.Entry(gen_frame, textvariable=self.title_var, width=28).pack(side="left", padx=(5, 10))
        self.rev_var = tk.StringVar(value="00")
        ttk.Label(gen_frame, text="Rev:").pack(side="left")
        ttk.Entry(gen_frame, textvariable=self.rev_var, width=5).pack(side="left", padx=(5, 10))
        self.cust_ver_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(gen_frame, text="Genera versione cliente", variable=self.cust_ver_var).pack(side="left",
                                                                                                    padx=(10, 10))
        ttk.Button(gen_frame, text="⚙️ TUTTO", command=self.generate_all).pack(side="right", padx=2)
        ttk.Button(gen_frame, text="📝 Solo CP", command=self.generate_cp_only).pack(side="right", padx=2)
        ttk.Button(gen_frame, text="📊 Solo FMEA", command=self.generate_fmea_only).pack(side="right", padx=2)
        ttk.Button(gen_frame, text="🔲 Solo FC", command=self.generate_fc_only).pack(side="right", padx=2)

        # BINDING PER DESELEZIONARE LA TABELLA CLICCANDO FUORI
        self.root.bind_all("<Button-1>", self._clear_tree_selection)

    def _clear_tree_selection(self, event):
        """Deseleziona gli elementi della tabella se si clicca fuori da essa e fuori dai pulsanti, ignorando i pop-up."""
        try:
            # Se il click avviene su una finestra pop-up (Toplevel), ignora la deselezione
            if event.widget.winfo_toplevel() != self.root:
                return

            widget_class = event.widget.winfo_class()
            # Classi di widget che devono mantenere la selezione attiva
            keep_selection_classes = ("Treeview", "TScrollbar", "Scrollbar", "TButton", "Button")

            if widget_class not in keep_selection_classes:
                if self.tree.selection():
                    self.tree.selection_remove(self.tree.selection())
        except Exception:
            pass

    def _add_file_row(self, frame, label, var, key, ext, row):
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(frame, textvariable=var, state="readonly", width=80).grid(row=row, column=1, sticky="we", padx=5,
                                                                            pady=2)
        ttk.Button(frame, text="Sfoglia...", command=lambda: self._choose_file(key, var, ext)).grid(row=row, column=2,
                                                                                                    sticky="w", padx=5,
                                                                                                    pady=2)

    def _get_master_sheet_paths(self, z, target_name):
        NS_WB = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        NS_RELS = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}

        wb_root = ET.fromstring(z.read('xl/workbook.xml'))
        sheet_rId = None
        for s in wb_root.findall('.//main:sheet', NS_WB):
            if target_name.lower() in s.get('name', '').lower():
                sheet_rId = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                break
        if not sheet_rId:
            sheet_rId = wb_root.find('.//main:sheet', NS_WB).get(
                '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')

        rels_wb = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        sheet_path = ""
        for r in rels_wb.findall('.//rel:Relationship', NS_RELS):
            if r.get('Id') == sheet_rId:
                sheet_path = f"xl/{r.get('Target')}"
                break

        drawing_path = None
        try:
            sh_xml = z.read(sheet_path)
            sh_root = ET.fromstring(sh_xml)
            dr_tag = sh_root.find('.//main:drawing', NS_WB)
            if dr_tag is not None:
                dr_rId = dr_tag.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                sh_rels_path = f"xl/worksheets/_rels/{os.path.basename(sheet_path)}.rels"
                if sh_rels_path in z.namelist():
                    rels_sh = ET.fromstring(z.read(sh_rels_path))
                    for r in rels_sh.findall('.//rel:Relationship', NS_RELS):
                        if r.get('Id') == dr_rId:
                            target = r.get('Target')
                            drawing_path = os.path.normpath(f"xl/worksheets/{target}").replace('\\', '/')
        except:
            pass
        return sheet_path, drawing_path

    def import_from_fc(self):
        filepath = filedialog.askopenfilename(title="Seleziona FC o MASTER",
                                              filetypes=[("Excel", "*.xlsm;*.xlsx"), ("Tutti", "*.*")])
        if not filepath: return
        self.show_loading("Analisi file in corso...")
        try:
            NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            data = {"title": "", "phases": []}
            with zipfile.ZipFile(filepath, 'r') as z:
                sh_path, dr_path = self._get_master_sheet_paths(z, "Flow Chart")

                # --- RISOLUZIONE DELLE SHARED STRINGS (Per file passati da Excel) ---
                shared_strings = []
                try:
                    sst_xml = z.read('xl/sharedStrings.xml')
                    sst_root = ET.fromstring(sst_xml)
                    for si in sst_root.findall('.//main:si', NS):
                        # Le stringhe possono avere formattazioni multiple dentro <r>, prendiamo tutto il testo
                        texts = [t.text for t in si.findall('.//main:t', NS) if t.text]
                        shared_strings.append("".join(texts))
                except KeyError:
                    pass  # sharedStrings.xml non esiste (il file ha usato inlineStr)

                def get_cell_value(c_elem):
                    """Estrae il valore da una cella, risolvendo le sharedStrings se necessario."""
                    if c_elem is None: return ""
                    t_attr = c_elem.get("t")
                    v_elem = c_elem.find("main:v", NS)
                    if t_attr == "s" and v_elem is not None:
                        try:
                            return shared_strings[int(v_elem.text)]
                        except:
                            return ""
                    elif t_attr == "inlineStr":
                        t_elem = c_elem.find("main:is/main:t", NS)
                        return t_elem.text if t_elem is not None else ""
                    elif v_elem is not None:
                        return v_elem.text
                    return ""

                # ---------------------------------------------------------------------

                root_sheet = ET.fromstring(z.read(sh_path))
                i = 1
                while True:
                    r_idx = 5 + 3 * i
                    row = root_sheet.find(f'.//main:row[@r="{r_idx}"]', NS)
                    if row is None: break

                    c_num = row.find(f'main:c[@r="B{r_idx}"]', NS)
                    c_ph = row.find(f'main:c[@r="D{r_idx}"]', NS)
                    c_ei = row.find(f'main:c[@r="F{r_idx}"]', NS)
                    c_dep = row.find(f'main:c[@r="H{r_idx}"]', NS)

                    val_num = get_cell_value(c_num)
                    if val_num:
                        name = get_cell_value(c_ph).replace(" *", "")
                        data["phases"].append({
                            "number": int(float(val_num)),
                            "phase": name,
                            "ext_int": get_cell_value(c_ei),
                            "department": get_cell_value(c_dep),
                            "note": ""
                        })
                    i += 1

                if dr_path and dr_path in z.namelist():
                    root_dr = ET.fromstring(z.read(dr_path))
                    for sp in root_dr.findall('.//xdr:sp', NS):
                        nv = sp.find('./xdr:nvSpPr/xdr:cNvPr', NS)
                        if nv is not None:
                            sp_name, sp_id = nv.get('name', ''), nv.get('id', '')
                            texts = [t.text for t in sp.findall('.//a:t', NS) if t.text]
                            if sp_id == "110" and len(texts) >= 2:
                                data["title"] = texts[1]
                            elif "Note" in sp_name:
                                for nt in texts[1:]:
                                    m = re.match(r'^\* Op\. (\d+)', nt)
                                    if m:
                                        op = int(m.group(1))
                                        txt = nt.split("):", 1)[-1].strip() if "):" in nt else nt
                                        for p in data["phases"]:
                                            if p["number"] == op: p["note"] = txt; break

            if data["phases"]:
                if self.phases_data and not messagebox.askyesno("Conferma", "Sovrascrivere le fasi attuali?"):
                    pass
                else:
                    self.phases_data.clear()
                    for item in self.tree.get_children(): self.tree.delete(item)
                    for p in data["phases"]:
                        m = re.match(r'^(.*?)\s*\((.*?)\)$', p["phase"])
                        base, comp = (m.group(1), m.group(2)) if m else (p["phase"], "")
                        is_cust = "RAW MATERIAL" in base.upper()
                        p_dict = {"number": p["number"], "phase": p["phase"], "base_phase": base, "component": comp,
                                  "ext_int": p["ext_int"], "department": p["department"], "note": p["note"],
                                  "is_customer": is_cust}
                        self.phases_data.append(p_dict)
                        self.tree.insert("", "end",
                                         values=(p["number"], p["phase"], p["ext_int"], p["department"], p["note"],
                                                 "Sì" if is_cust else "No"))
                    if data["title"]: self.title_var.set(data["title"])
                    self._renumber()
                    messagebox.showinfo("Fatto", f"Importate {len(data['phases'])} fasi.")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
        finally:
            self.hide_loading()

    def _choose_db_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]);
        if f: self.real_paths["db_master"] = f; self.db_master_var.set(
            format_short_path(f)); self.archive = self._load_archive_from_excel(); self._refresh_archive_combo();
        self.saved_config["db_master"] = f;
        save_config(self.saved_config)

    def _choose_db_cp_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]);
        if f: self.real_paths["db_master_cp"] = f; self.db_master_var_cp.set(format_short_path(f)); self.saved_config[
            "db_master_cp"] = f; save_config(self.saved_config)

    def _choose_file(self, k, var, ext):
        f = filedialog.askopenfilename(filetypes=[("Excel", f"*{ext}")]);
        if f: self.real_paths[k] = f; var.set(format_short_path(f)); self.saved_config[k] = f; save_config(
            self.saved_config)

    def _load_archive_from_excel(self):
        try:
            wb = openpyxl.load_workbook(self.real_paths["db_master"], data_only=True)
            ws = wb["Fasi"] if "Fasi" in wb.sheetnames else wb.create_sheet("Fasi")
            arc = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r and r[0]: arc.append(
                    {"phase": str(r[0]), "ext_int": str(r[1] or "Internal"), "department": str(r[2] or "")})
            self._ensure_phases_in_db(wb, arc, self.real_paths["db_master"]);
            wb.close();
            return arc
        except:
            return []

    def _ensure_phases_in_db(self, wb, arc, path):
        m_ws = wb["MASTERS"] if "MASTERS" in wb.sheetnames else wb.create_sheet("MASTERS")
        f_ws = wb["database_funzionale"] if "database_funzionale" in wb.sheetnames else wb.create_sheet(
            "database_funzionale")
        exist = {str(r[0]).strip().lower() for r in m_ws.iter_rows(min_row=1, max_col=1, values_only=True) if
                 r and r[0]}
        mod = False
        for p in arc:
            if p["phase"].strip().lower() not in exist: m_ws.append([p["phase"]]); f_ws.append([p["phase"]]); mod = True
        if mod:
            try:
                wb.save(path);
                format_masters_db(path)
            except:
                pass

    def _refresh_archive_combo(self):
        self.phase_combo["values"] = [e["phase"] for e in self.archive]
        self.dept_combo["values"] = list({e["department"] for e in self.archive if e["department"]})

    def _on_phase_select(self, e=None):
        sel = self.phase_var.get().strip().lower()
        for p in self.archive:
            if p["phase"].strip().lower() == sel:
                self.extint_var.set(p["ext_int"])
                self.dept_var.set(p["department"])
                break

    def _autocomplete(self, e, combo, full_list):
        keysym = e.keysym

        # 1. Ignora i tasti modificatori e quelli di navigazione che causano bug
        ignore_keys = (
            "Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R",
            "Caps_Lock", "Return", "Tab", "Escape", "Left", "Right", "Up", "Down", "Home", "End"
        )
        if keysym in ignore_keys:
            return

        txt = combo.get()
        if not txt:
            combo["values"] = full_list
            return

        # 2. Aggiorna sempre la lista suggerimenti a tendina
        match = [i for i in full_list if txt.lower() in i.lower()]
        combo["values"] = match

        # 3. Se l'utente sta cancellando, fermati (non sovrascrivere forzatamente)
        if keysym in ("BackSpace", "Delete"):
            return

        # 4. Scrive ed evidenzia il miglior risultato
        for i in match:
            if i.lower().startswith(txt.lower()):
                combo.set(i)
                combo.icursor(len(txt))
                combo.selection_range(len(txt), 'end')
                # 5. INVIA IL SEGNALE DI CAMBIAMENTO (Autocompleta subito Reparto e Int/Ext)
                combo.event_generate("<<ComboboxSelected>>")
                break

    def _get_unique_depts(self):
        return list({e["department"] for e in self.archive if e["department"]})

    def add_phase(self, e=None):
        n, ph, ei, dp, co, nt = self.num_var.get(), self.phase_var.get(), self.extint_var.get(), self.dept_var.get(), self.comp_var.get(), self.note_var.get()

        if not all([n, ph, ei, dp]):
            messagebox.showwarning("Mancano dati",
                                   "Compila i campi obbligatori:\n- N° Op.\n- Nome Fase\n- Int / Ext\n- Reparto")
            return

        try:
            num = int(n)
        except:
            return

        disp = f"{ph} ({co})" if co else ph
        cust = "RAW MATERIAL" in ph.upper()

        d = {
            "number": num, "phase": disp, "base_phase": ph, "component": co,
            "ext_int": ei, "department": dp, "note": nt, "is_customer": cust
        }

        self.phases_data.append(d)
        self.tree.insert("", "end", values=(num, disp, ei, dp, nt, "Sì" if cust else "No"))

        # ── SVUOTA TUTTI I CAMPI ──
        self.num_var.set(str(num + 10))
        self.phase_var.set("")
        self.comp_var.set("")
        self.extint_var.set("Internal")
        self.dept_var.set("")
        self.note_var.set("")
        # ──────────────────────────

    def _renumber(self):
        for i, p in enumerate(self.phases_data): p["number"] = (i + 1) * 10
        for i, item in enumerate(self.tree.get_children()):
            v = list(self.tree.item(item, "values"))
            v[0] = self.phases_data[i]["number"]
            self.tree.item(item, values=v)
        self.num_var.set(str((len(self.phases_data) + 1) * 10))

    def clear_table(self):
        if not self.phases_data: return
        if messagebox.askyesno("Conferma", "Vuoi davvero rimuovere TUTTE le fasi inserite?"):
            self.phases_data.clear()
            for item in self.tree.get_children(): self.tree.delete(item)
            self._renumber()

    def remove_phase(self):
        selected_items = self.tree.selection()
        if not selected_items: return
        indices = sorted([self.tree.index(item) for item in selected_items], reverse=True)
        for idx in indices: self.phases_data.pop(idx)
        for item in selected_items: self.tree.delete(item)
        self._renumber()

    def move_up(self):
        selected = self.tree.selection()
        if not selected: return
        if len(selected) > 1: messagebox.showinfo("Avviso",
                                                  "Seleziona una sola fase alla volta per poterla spostare."); return
        item = selected[0]
        idx = self.tree.index(item)
        if idx > 0:
            self.phases_data[idx - 1], self.phases_data[idx] = self.phases_data[idx], self.phases_data[idx - 1]
            self.tree.move(item, "", idx - 1);
            self._renumber()

    def move_down(self):
        selected = self.tree.selection()
        if not selected: return
        if len(selected) > 1: messagebox.showinfo("Avviso",
                                                  "Seleziona una sola fase alla volta per poterla spostare."); return
        item = selected[0]
        idx = self.tree.index(item)
        if idx < len(self.phases_data) - 1:
            self.phases_data[idx + 1], self.phases_data[idx] = self.phases_data[idx], self.phases_data[idx + 1]
            self.tree.move(item, "", idx + 1);
            self._renumber()

    def _on_double_click(self, e):
        item = self.tree.identify_row(e.y)
        if item: self._open_edit_dialog(self.tree.index(item), item)

    def _open_edit_dialog(self, idx, item):
        phase_dict = self.phases_data[idx]
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Modifica fase — N° {phase_dict['number']}")
        dialog.resizable(False, False);
        dialog.grab_set();
        dialog.transient(self.root)

        dialog.update_idletasks()
        dw, dh = dialog.winfo_reqwidth(), dialog.winfo_reqheight()
        sw, sh = dialog.winfo_screenwidth(), dialog.winfo_screenheight()
        dialog.geometry(f"+{int(sw / 2 - dw / 2)}+{int(sh / 2 - dh / 2)}")

        pad = {"padx": 10, "pady": 6}

        ttk.Label(dialog, text="N° Operazione:").grid(row=0, column=0, sticky="e", **pad)
        num_var = tk.StringVar(value=str(phase_dict["number"]))
        ttk.Entry(dialog, textvariable=num_var, width=10).grid(row=0, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Nome Fase:").grid(row=1, column=0, sticky="e", **pad)
        phase_var = tk.StringVar(value=phase_dict.get("base_phase", phase_dict["phase"]))
        phase_combo = ttk.Combobox(dialog, textvariable=phase_var, values=[p["phase"] for p in self.archive], width=50)
        phase_combo.grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Componente:").grid(row=2, column=0, sticky="e", **pad)
        comp_var = tk.StringVar(value=phase_dict.get("component", ""))
        ttk.Entry(dialog, textvariable=comp_var, width=25).grid(row=2, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Int / Ext:").grid(row=3, column=0, sticky="e", **pad)
        extint_var = tk.StringVar(value=phase_dict["ext_int"])
        ttk.Combobox(dialog, textvariable=extint_var, values=["Internal", "External"], width=12, state="readonly").grid(
            row=3, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Reparto:").grid(row=4, column=0, sticky="e", **pad)
        dept_var = tk.StringVar(value=phase_dict["department"])
        dept_combo = ttk.Combobox(dialog, textvariable=dept_var, values=self._get_unique_depts(), width=50)
        dept_combo.grid(row=4, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Additional remarks:").grid(row=5, column=0, sticky="e", **pad)
        note_var = tk.StringVar(value=phase_dict.get("note", ""))
        ttk.Entry(dialog, textvariable=note_var, width=52).grid(row=5, column=1, sticky="w", **pad)

        ttk.Label(dialog, text="Includere in Versione Cliente:").grid(row=6, column=0, sticky="e", **pad)
        cust_var = tk.BooleanVar(value=phase_dict.get("is_customer", False))
        ttk.Checkbutton(dialog, variable=cust_var).grid(row=6, column=1, sticky="w", **pad)

        def _on_edit_phase_select(e=None):
            sel = phase_var.get().strip().lower()
            for p in self.archive:
                if p["phase"].strip().lower() == sel:
                    extint_var.set(p["ext_int"]);
                    dept_var.set(p["department"]);
                    phase_var.set(p["phase"]);
                    break

        phase_combo.bind("<<ComboboxSelected>>", _on_edit_phase_select)
        phase_combo.bind("<FocusOut>", _on_edit_phase_select)
        phase_combo.bind("<KeyRelease>",
                         lambda e: self._autocomplete(e, phase_combo, [p["phase"] for p in self.archive]))
        dept_combo.bind("<KeyRelease>", lambda e: self._autocomplete(e, dept_combo, self._get_unique_depts()))

        def _confirm():
            comp = comp_var.get().strip()
            base_ph = phase_var.get().strip()
            display_phase = f"{base_ph} ({comp})" if comp else base_ph

            self.phases_data[idx]["number"] = int(num_var.get())
            self.phases_data[idx]["phase"] = display_phase
            self.phases_data[idx]["base_phase"] = base_ph
            self.phases_data[idx]["component"] = comp
            self.phases_data[idx]["ext_int"] = extint_var.get().strip()
            self.phases_data[idx]["department"] = dept_var.get().strip()
            self.phases_data[idx]["note"] = note_var.get().strip()
            self.phases_data[idx]["is_customer"] = cust_var.get()

            self.tree.item(item, values=(num_var.get(), display_phase, extint_var.get().strip(), dept_var.get().strip(),
                                         note_var.get().strip(), "Sì" if cust_var.get() else "No"))
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.grid(row=7, column=0, columnspan=2, pady=(4, 10))
        ttk.Button(btn_frame, text="\u2705 Conferma", command=_confirm).pack(side="left", padx=8)

    def _save_to_archive(self):
        ph, ei, dp = self.phase_var.get(), self.extint_var.get(), self.dept_var.get()
        if not ph or not dp: return
        for e in self.archive:
            if e["phase"].lower() == ph.lower(): e.update(
                {"ext_int": ei, "department": dp}); self._save_archive_to_excel(); self._refresh_archive_combo(); return
        self.archive.append({"phase": ph, "ext_int": ei, "department": dp});
        self._save_archive_to_excel();
        self._refresh_archive_combo()

    def _delete_from_archive(self):
        ph = self.phase_var.get()
        if ph and messagebox.askyesno("Conferma", "Eliminare?"):
            self.archive = [e for e in self.archive if e["phase"].lower() != ph.lower()]
            self._save_archive_to_excel();
            self._refresh_archive_combo()

    def _save_archive_to_excel(self):
        try:
            wb = openpyxl.load_workbook(self.real_paths["db_master"])
            ws = wb["Fasi"]
            ws.delete_rows(2, ws.max_row)
            for p in self.archive: ws.append([p["phase"], p["ext_int"], p["department"]])
            self._ensure_phases_in_db(wb, self.archive, self.real_paths["db_master"])
        except:
            pass

    def format_master_db(self):
        try:
            format_masters_db(self.real_paths["db_master"]);
            messagebox.showinfo("OK", "DB Formattato")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def format_master_cp_db_gui(self):
        try:
            format_master_cp_db(self.real_paths["db_master_cp"]);
            messagebox.showinfo("OK", "DB CP Formattato")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def show_loading(self, msg):
        self.lw = tk.Toplevel(self.root);
        self.lw.title("Attendere prego");
        self.lw.geometry("380x120")
        self.lw.resizable(False, False);
        self.lw.transient(self.root);
        self.lw.grab_set()
        self.lw.protocol("WM_DELETE_WINDOW", lambda: None)
        self.lw.update_idletasks()
        x, y = int(self.lw.winfo_screenwidth() / 2 - 190), int(self.lw.winfo_screenheight() / 2 - 60)
        self.lw.geometry(f"+{x}+{y}")
        ttk.Label(self.lw, text="⏳", font=("Segoe UI", 26)).pack(pady=(10, 0))
        ttk.Label(self.lw, text=msg, justify="center", font=("Segoe UI", 11, "bold")).pack(pady=(5, 10))
        self.root.update()

    def hide_loading(self):
        if hasattr(self, 'lw') and self.lw.winfo_exists():
            self.lw.grab_release();
            self.lw.destroy()

    def _check_pre_gen(self):
        if not self.phases_data:
            messagebox.showwarning("Attenzione", "Aggiungi almeno una fase al progetto!")
            return None
        title = self.title_var.get().strip()
        if title == "NOME PRODOTTO":
            if not messagebox.askyesno("Conferma Nome",
                                       "Il nome prodotto è 'NOME PRODOTTO'. Sicuro di voler continuare?"):
                return None
        return title

    def _create_output_folder(self, title):
        folder_name = f"{title} Pratiche"
        os.makedirs(folder_name, exist_ok=True)
        return folder_name

    def generate_fc_only(self):
        title = self._check_pre_gen()
        if not title: return
        template_name = self.real_paths["tpl_fc"]
        folder = self._create_output_folder(title)
        output_name = os.path.join(folder, f"{title}_FC.xlsm")

        if not os.path.exists(template_name):
            messagebox.showerror("Errore", "Verifica che il Template Flow Chart esista.")
            return

        self.show_loading("Generazione Flow Chart in corso...")
        try:
            generate_fc(template_name, output_name, self.phases_data, title=title)
            self.hide_loading()
            messagebox.showinfo("Successo", f"Flow Chart generato in:\n{output_name}")
        except Exception as e:
            self.hide_loading();
            messagebox.showerror("Errore", str(e))

    def generate_fmea_only(self):
        title = self._check_pre_gen()
        if not title: return
        rev, gen_cust = self.rev_var.get().strip(), self.cust_ver_var.get()
        tpl, db = self.real_paths["tpl_fmea"], self.real_paths["db_master"]
        folder = self._create_output_folder(title)
        out = os.path.join(folder, f"{title}_FMEA.xlsx")
        temp_main = os.path.join(folder, f"temp_fmea_{title}.xlsx")
        temp_cust = os.path.join(folder, f"temp_fmea_cust_{title}.xlsx")

        if not os.path.exists(tpl) or not os.path.exists(db):
            messagebox.showerror("Errore", "Verifica che Template e Database esistano.")
            return

        self.show_loading("Generazione FMEA in corso...")
        try:
            if gen_cust:
                import win32com.client as win32
                generate_fmea(tpl, db, temp_main, self.phases_data, title=title, rev=rev, customer_only=False)
                generate_fmea(tpl, db, temp_cust, self.phases_data, title=title, rev=rev, customer_only=True)
                excel = win32.DispatchEx('Excel.Application')
                excel.DisplayAlerts = False;
                excel.Visible = False
                try:
                    wb_main = excel.Workbooks.Open(os.path.abspath(temp_main))
                    if os.path.exists(temp_cust):
                        wb_cust = excel.Workbooks.Open(os.path.abspath(temp_cust))
                        wb_cust.Sheets(1).Copy(None, wb_main.Sheets(wb_main.Sheets.Count))
                        wb_cust.Close(False)
                    wb_main.SaveAs(os.path.abspath(out), FileFormat=51)
                    wb_main.Close(False)
                finally:
                    excel.Quit()
                for f in [temp_main, temp_cust]:
                    if os.path.exists(f): os.remove(f)
                self.hide_loading();
                messagebox.showinfo("Successo", f"FMEA (con versione Cliente) generato in:\n{out}")
            else:
                generate_fmea(tpl, db, out, self.phases_data, title=title, rev=rev, customer_only=False)
                self.hide_loading();
                messagebox.showinfo("Successo", f"FMEA generato in:\n{out}")
        except Exception as e:
            self.hide_loading();
            messagebox.showerror("Errore", f"Dettaglio: {str(e)}")

    def generate_cp_only(self):
        title = self._check_pre_gen()
        if not title: return
        db, rev, gen_cust, tpl = self.real_paths["db_master_cp"], self.rev_var.get().strip(), self.cust_ver_var.get(), \
            self.real_paths["tpl_cp"]
        folder = self._create_output_folder(title)
        out = os.path.join(folder, f"{title}_CP.xlsx")
        temp_main = os.path.join(folder, f"temp_cp_{title}.xlsx")
        temp_cust = os.path.join(folder, f"temp_cp_cust_{title}.xlsx")

        if not os.path.exists(tpl):
            messagebox.showerror("Errore", "Verifica che il Template Control Plan esista.")
            return

        self.show_loading("Elaborazione e Analisi Control Plan...")
        try:
            cache = {}
            if gen_cust:
                import win32com.client as win32
                generate_cp(tpl, db, temp_main, self.phases_data, title=title, rev=rev, customer_only=False,
                            scelte_cache=cache)
                self.lw.title("Unione file...");
                self.root.update()
                generate_cp(tpl, db, temp_cust, self.phases_data, title=title, rev=rev, customer_only=True,
                            scelte_cache=cache)
                excel = win32.DispatchEx('Excel.Application')
                excel.DisplayAlerts = False;
                excel.Visible = False
                try:
                    wb_main = excel.Workbooks.Open(os.path.abspath(temp_main))
                    if os.path.exists(temp_cust):
                        wb_cust = excel.Workbooks.Open(os.path.abspath(temp_cust))
                        wb_cust.Sheets(1).Copy(None, wb_main.Sheets(wb_main.Sheets.Count))
                        wb_cust.Close(False)
                    wb_main.SaveAs(os.path.abspath(out), FileFormat=51)
                    wb_main.Close(False)
                finally:
                    excel.Quit()
                for f in [temp_main, temp_cust]:
                    if os.path.exists(f): os.remove(f)
                self.hide_loading();
                messagebox.showinfo("Successo", f"Control Plan generato in:\n{out}")
            else:
                generate_cp(tpl, db, out, self.phases_data, title=title, rev=rev, customer_only=False,
                            scelte_cache=cache)
                self.hide_loading();
                messagebox.showinfo("Successo", f"Control Plan generato in:\n{out}")
        except Exception as e:
            self.hide_loading()
            if "annullata dall'utente" not in str(e): messagebox.showerror("Errore", f"Dettaglio: {str(e)}")

    def generate_all(self):
        title = self._check_pre_gen()
        if not title: return
        rev, gen_cust = self.rev_var.get().strip(), self.cust_ver_var.get()
        folder = self._create_output_folder(title)
        out = os.path.join(folder, f"{title}_MASTER.xlsm")

        temps = {
            "fc": os.path.join(folder, f"temp_fc_{title}.xlsm"),
            "fmea": os.path.join(folder, f"temp_fmea_{title}.xlsx"),
            "fmea_c": os.path.join(folder, f"temp_fmea_cust_{title}.xlsx"),
            "cp": os.path.join(folder, f"temp_cp_{title}.xlsx"),
            "cp_c": os.path.join(folder, f"temp_cp_cust_{title}.xlsx")
        }

        self.show_loading("Elaborazione dati e interrogazione Database...")
        try:
            import win32com.client as win32
            cache = {}
            generate_fc(self.real_paths["tpl_fc"], temps["fc"], self.phases_data, title=title)
            generate_fmea(self.real_paths["tpl_fmea"], self.real_paths["db_master"], temps["fmea"], self.phases_data,
                          title=title, rev=rev, customer_only=False)
            generate_cp(self.real_paths["tpl_cp"], self.real_paths["db_master_cp"], temps["cp"], self.phases_data,
                        title=title, rev=rev, customer_only=False, scelte_cache=cache)

            if gen_cust:
                generate_fmea(self.real_paths["tpl_fmea"], self.real_paths["db_master"], temps["fmea_c"],
                              self.phases_data, title=title, rev=rev, customer_only=True)
                generate_cp(self.real_paths["tpl_cp"], self.real_paths["db_master_cp"], temps["cp_c"], self.phases_data,
                            title=title, rev=rev, customer_only=True, scelte_cache=cache)

            self.lw.title("Unione file")
            for w in self.lw.winfo_children():
                if isinstance(w, ttk.Label) and "Elaborazione" in w.cget("text"):
                    w.config(text="Fase finale: unione dei documenti in corso...\nNon chiudere Excel.")
            self.root.update()

            excel = win32.DispatchEx('Excel.Application')
            excel.DisplayAlerts = False;
            excel.Visible = False
            try:
                wb_fc = excel.Workbooks.Open(os.path.abspath(temps["fc"]))
                for key in ["fmea", "fmea_c", "cp", "cp_c"]:
                    if os.path.exists(temps[key]):
                        wb_src = excel.Workbooks.Open(os.path.abspath(temps[key]))
                        wb_src.Sheets(1).Copy(None, wb_fc.Sheets(wb_fc.Sheets.Count))
                        wb_src.Close(False)
                wb_fc.SaveAs(os.path.abspath(out), FileFormat=52)
                wb_fc.Close(False)
            finally:
                excel.Quit()

            for f in temps.values():
                if os.path.exists(f): os.remove(f)
            self.hide_loading();
            messagebox.showinfo("Successo", f"File MASTER completo creato in:\n{out}")
        except Exception as e:
            self.hide_loading()
            if "annullata" not in str(e): messagebox.showerror("Errore", str(e))
            for f in temps.values():
                try:
                    os.remove(f)
                except:
                    pass


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    try:
        root.iconbitmap(resource_path("logo.ico"))
    except:
        pass
    app = FlowChartApp(root)
    root.deiconify()
    root.mainloop()